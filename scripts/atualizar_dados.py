#!/usr/bin/env python3
import datetime as dt
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SOURCE_FILE = Path(os.environ.get("SOURCE_FILE", "Radar_Brasil_2027.xlsx"))
OUTPUT = Path(os.environ.get("OUTPUT_FILE", "dados.json"))
NEWS_OUTPUT = Path(os.environ.get("NEWS_OUTPUT_FILE", "noticias.json"))
SHEET = "02_Eventos"
NEWS_SHEET = "06_Noticias"
FIELDS = [
    "ID", "Titulo", "Status", "Data", "DataBR", "UF", "Cidade", "Categoria",
    "Organizador", "Publico", "Patrocinador", "Local", "Latitude",
    "Longitude", "Link", "Observacoes", "Mes", "Ano", "Regiao",
]
MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
          "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
REGIONS = {
    "AC": "Norte", "AP": "Norte", "AM": "Norte", "PA": "Norte",
    "RO": "Norte", "RR": "Norte", "TO": "Norte",
    "AL": "Nordeste", "BA": "Nordeste", "CE": "Nordeste",
    "MA": "Nordeste", "PB": "Nordeste", "PE": "Nordeste",
    "PI": "Nordeste", "RN": "Nordeste", "SE": "Nordeste",
    "DF": "Centro-Oeste", "GO": "Centro-Oeste",
    "MT": "Centro-Oeste", "MS": "Centro-Oeste",
    "ES": "Sudeste", "MG": "Sudeste", "RJ": "Sudeste", "SP": "Sudeste",
    "PR": "Sul", "RS": "Sul", "SC": "Sul", "BR": "Nacional",
}
SOURCE_REQUIRED = [
    "ID", "Status", "Data", "UF", "Cidade", "Categoria", "Organizador",
    "Publico", "Patrocinador", "Local", "Latitude", "Longitude",
    "Link", "Observacoes",
]


def normalized(value):
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", text.lower())


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Data inválida: {value!r}")


def as_number(value, default=None):
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, (int, float)):
        number = value
    else:
        text = str(value).strip().replace(".", "").replace(",", ".")
        number = float(text)
    return int(number) if float(number).is_integer() else float(number)


def fallback_title(event):
    """Cria um título legível quando a planilha ainda não tiver a coluna Titulo."""
    observation = str(event.get("Observacoes") or "").strip()
    first_sentence = re.split(r"(?<=[.!?])\s+", observation, maxsplit=1)[0]
    first_sentence = first_sentence.rstrip(". ")
    if first_sentence:
        return first_sentence[:140].rstrip()
    category = str(event.get("Categoria") or "Evento").strip()
    city = str(event.get("Cidade") or "").strip()
    return f"{category} — {city}" if city else category


def main():
    if not SOURCE_FILE.is_file():
        raise RuntimeError(f"A planilha {SOURCE_FILE} não foi encontrada no repositório.")
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise RuntimeError(f"A aba {SHEET!r} não foi encontrada.")
    sheet = workbook[SHEET]

    wanted = {normalized(field): field for field in FIELDS}
    wanted.update({
        "publicoestimado": "Publico",
        "linkfonte": "Link",
    })
    header_row = None
    columns = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        found = {wanted[normalized(value)]: index for index, value in enumerate(row)
                 if normalized(value) in wanted}
        if all(name in found for name in SOURCE_REQUIRED):
            header_row, columns = row_number, found
            break
    if header_row is None:
        raise RuntimeError("Cabeçalho da aba 02_Eventos não reconhecido.")
    missing = [field for field in SOURCE_REQUIRED if field not in columns]
    if missing:
        raise RuntimeError("Campos obrigatórios ausentes: " + ", ".join(missing))

    existing_titles = {}
    if OUTPUT.is_file():
        try:
            existing_titles = {
                str(item.get("ID", "")): str(item.get("Titulo", "")).strip()
                for item in json.loads(OUTPUT.read_text(encoding="utf-8"))
                if item.get("ID") and item.get("Titulo")
            }
        except (OSError, ValueError, TypeError):
            existing_titles = {}

    events = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        event_id = row[columns["ID"]]
        if event_id is None or not str(event_id).strip():
            continue
        event = {
            field: row[columns[field]] if field in columns else None
            for field in FIELDS
        }
        date = as_date(event["Data"])
        for field in ("ID", "Status", "UF", "Cidade", "Categoria", "Organizador",
                      "Titulo", "Patrocinador", "Local", "Link", "Observacoes", "Regiao"):
            event[field] = "" if event[field] is None else str(event[field]).strip()
        if not event["Titulo"]:
            event["Titulo"] = existing_titles.get(event["ID"]) or fallback_title(event)
        event["Data"] = date.isoformat()
        event["DataBR"] = date.strftime("%d/%m/%Y")
        event["Publico"] = as_number(event["Publico"], 0)
        event["Latitude"] = as_number(event["Latitude"])
        event["Longitude"] = as_number(event["Longitude"])
        event["Mes"] = MONTHS[date.month - 1]
        event["Ano"] = date.year
        event["Regiao"] = REGIONS.get(event["UF"].upper(), "")
        events.append({field: event[field] for field in FIELDS})

    if not events:
        raise RuntimeError("Nenhum evento válido foi encontrado; dados.json não será alterado.")
    ids = [event["ID"] for event in events]
    if len(ids) != len(set(ids)):
        raise RuntimeError("Há IDs de eventos duplicados; dados.json não será alterado.")

    OUTPUT.write_text(json.dumps(events, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if NEWS_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"A aba {NEWS_SHEET!r} não foi encontrada.")
    news_sheet = workbook[NEWS_SHEET]
    news_wanted = {
        "data": "Data", "titulo": "Titulo", "tema": "Tema",
        "cidadeuf": "CidadeUF", "veiculo": "Veiculo", "link": "Link",
        "sentimento": "Sentimento", "impacto": "Impacto", "resumo": "Resumo",
    }
    news_header = None
    news_columns = {}
    for row_number, row in enumerate(
        news_sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1
    ):
        found = {
            news_wanted[normalized(value)]: index
            for index, value in enumerate(row)
            if normalized(value) in news_wanted
        }
        if all(field in found for field in news_wanted.values()):
            news_header, news_columns = row_number, found
            break
    if news_header is None:
        raise RuntimeError("Cabeçalho da aba 06_Noticias não reconhecido.")

    news = []
    for row in news_sheet.iter_rows(min_row=news_header + 1, values_only=True):
        title = row[news_columns["Titulo"]]
        if title is None or not str(title).strip():
            continue
        date = as_date(row[news_columns["Data"]])
        item = {
            "Data": date.isoformat(),
            "DataBR": date.strftime("%d/%m/%Y"),
            "Titulo": str(title).strip(),
            "Tema": str(row[news_columns["Tema"]] or "").strip(),
            "CidadeUF": str(row[news_columns["CidadeUF"]] or "").strip(),
            "Veiculo": str(row[news_columns["Veiculo"]] or "").strip(),
            "Link": str(row[news_columns["Link"]] or "").strip(),
            "Sentimento": str(row[news_columns["Sentimento"]] or "").strip(),
            "Impacto": str(row[news_columns["Impacto"]] or "").strip(),
            "Resumo": str(row[news_columns["Resumo"]] or "").strip(),
        }
        news.append(item)
    news.sort(key=lambda item: item["Data"], reverse=True)
    NEWS_OUTPUT.write_text(
        json.dumps(news, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"{len(events)} eventos gravados em {OUTPUT}; "
        f"{len(news)} notícias gravadas em {NEWS_OUTPUT}"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        raise
